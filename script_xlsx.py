from openpyxl import load_workbook

# Загрузка книги Excel
workbook = load_workbook('tmp/New XLSX.xlsx')

# Получение активного листа
sheet = workbook.active

# 1. Получение значения из конкретной ячейки
print(f"Значение из ячейки A1: {sheet.cell(row=1, column=1).value}")

# 2. Запись значения в ячейку
sheet.cell(row=2, column=1).value = "Новая запись"  # Запись в ячейку A2
print("Записано 'Новая запись' в ячейку A2")

# 3. Получение значения через адрес ячейки
print(f"Значение из ячейки B2: {sheet['B2'].value}")

# 4. Запись значения через адрес ячейки
sheet["C3"] = "OpenPyXL"  # Запись значения в ячейку C3
print("Записано 'OpenPyXL' в ячейку C3")

# 5. Итерация по строкам
print("Значения в первой строке:")
for cell in sheet[1]:  # Перебор всех ячеек первой строки
    print(f" - {cell.value}")

# 6. Итерация по столбцам
print("Значения в первом столбце:")
for cell in sheet['A']:  # Перебор всех ячеек столбца A
    print(f" - {cell.value}")

# 7. Получение всех строк
print("Все строки на листе (первые 5):")
for row in sheet.iter_rows(min_row=1, max_row=5, values_only=True):  # Строки 1-5
    print(f" - {row}")

# 8. Получение всех столбцов
print("Все столбцы на листе (первые 3):")
for col in sheet.iter_cols(min_col=1, max_col=3, values_only=True):  # Столбцы 1-3
    print(f" - {col}")

# 9. Получение размеров листа
print(f"Количество строк на листе: {sheet.max_row}")
print(f"Количество столбцов на листе: {sheet.max_column}")

# 10. Получение названия листа
print(f"Название текущего листа: {sheet.title}")

# Сохранение изменений в книге
workbook.save('tmp/Updated_New_XLSX.xlsx')
print("Изменения сохранены в 'tmp/Updated_New_XLSX.xlsx'")