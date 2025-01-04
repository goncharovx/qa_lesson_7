import csv
import os
import shutil
import zipfile

import pytest
from openpyxl import load_workbook
from pypdf import PdfReader


@pytest.fixture(scope="module")
def setup_resources():
    # Подготовка тестовых ресурсов перед выполнением тестов.
    resources_dir = "resources"  # Папка с ресурсами
    pdf_path = os.path.join(resources_dir, "Python-Testing.pdf")  # Путь к PDF файлу
    xlsx_path = os.path.join(resources_dir, "example.xlsx")  # Путь к XLSX файлу
    csv_path = os.path.join(resources_dir, "example.csv")  # Путь к CSV файлу
    zip_path = os.path.join(resources_dir, "hw_archive.zip")  # Путь к ZIP архиву

    # Проверяем существование папки ресурсов
    assert os.path.exists(resources_dir), f"Папка '{resources_dir}' не существует."

    # Проверяем существование PDF файла
    assert os.path.exists(pdf_path), f"PDF файл '{pdf_path}' не существует."

    # Создаём XLSX файл, если он отсутствует
    if not os.path.exists(xlsx_path):
        create_xlsx(xlsx_path)
    # Создаём CSV файл, если он отсутствует
    if not os.path.exists(csv_path):
        create_csv(csv_path)

    # Создаём архив, если он отсутствует
    if not os.path.exists(zip_path):
        with zipfile.ZipFile(zip_path, "w") as archive:
            archive.write(pdf_path, arcname="Python-Testing.pdf")  # Добавляем PDF в архив
            archive.write(csv_path, arcname="example.csv")  # Добавляем CSV в архив
            archive.write(xlsx_path, arcname="example.xlsx")  # Добавляем XLSX в архив

    # Передаём пути для использования в тестах
    yield {
        "resources_dir": resources_dir,
        "zip_path": zip_path,
        "xlsx_path": xlsx_path,
        "csv_path": csv_path
    }

    # Очистка ресурсов после выполнения всех тестов
    cleanup([xlsx_path, csv_path, zip_path], "extracted")


def create_xlsx(xlsx_path):
    # Создание XLSX файла.
    from openpyxl import Workbook
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    xlsx_data = [["name", "age", "city"], ["Anna", 25, "Saint-Petersburg"], ["Ivan", 30, "Moscow"]]
    for row in xlsx_data:
        sheet.append(row)  # Добавляем строки с данными
    workbook.save(xlsx_path)  # Сохраняем файл


def create_csv(csv_path):
    # Создание CSV файла.
    csv_data = [["name", "age", "city"], ["Petr", "35", "Kazan"], ["Elena", "28", "Novosibirsk"]]
    with open(csv_path, "w", newline="", encoding="utf-8") as csv_file:
        writer = csv.writer(csv_file)
        writer.writerows(csv_data)


def cleanup(files_to_remove, directory_to_remove):
    # Удаление временных файлов и папок.
    for file_path in files_to_remove:
        if os.path.exists(file_path):
            os.remove(file_path)

    if os.path.exists(directory_to_remove):
        shutil.rmtree(directory_to_remove)


def test_create_zip(setup_resources):
    # Тест создания ZIP архива.
    zip_path = setup_resources["zip_path"]
    assert os.path.exists(zip_path), f"ZIP архив '{zip_path}' не создан."


def test_verify_pdf_content(setup_resources):
    # Тест проверки содержимого PDF файла.=
    zip_path = setup_resources["zip_path"]

    with zipfile.ZipFile(zip_path, 'r') as archive:
        """
        Используется метод archive.open для чтения PDF файла.
        Подходит для потокового чтения больших файлов.
        Файл читается из архива без извлечения.
        """
        with archive.open("Python-Testing.pdf") as pdf_file:
            pdf_reader = PdfReader(pdf_file)
            page_text = pdf_reader.pages[10].extract_text()  # Читаем текст 11-й страницы
            expected_text = "Learn pytest While Testing an Example Application"
            assert expected_text in page_text, "Содержимое PDF файла некорректно."
            assert len(pdf_reader.pages) > 10, "PDF файл имеет меньше 11 страниц."


def test_verify_csv_content(setup_resources):
    # Тест проверки содержимого CSV файла.
    zip_path = setup_resources["zip_path"]

    with zipfile.ZipFile(zip_path, 'r') as archive:
        """
        Используется метод archive.read для чтения CSV файла.
        Подходит для небольших файлов, которые нужно быстро загрузить в память.
        Считывается весь файл целиком.
        """
        csv_content = archive.read("example.csv").decode("utf-8")  # Читаем CSV файл в память
        csv_rows = [row.split(",") for row in csv_content.splitlines()]  # Преобразуем содержимое в список строк
        expected_csv_data = [["name", "age", "city"], ["Petr", "35", "Kazan"], ["Elena", "28", "Novosibirsk"]]
        assert csv_rows == expected_csv_data, f"Содержимое CSV файла некорректно: {csv_rows} != {expected_csv_data}"
        assert len(csv_rows) > 2, "CSV файл содержит недостаточно строк."


def test_verify_xlsx_content(setup_resources):
    # Тест проверки содержимого XLSX файла.
    zip_path = setup_resources["zip_path"]

    with zipfile.ZipFile(zip_path, 'r') as archive:
        """
        Используется метод archive.extract для извлечения XLSX файла.
        Подходит для файлов, которые нужно обработать сторонними библиотеками.
        Извлекает файл из архива на диск.
        """
        extracted_xlsx_path = archive.extract("example.xlsx", path="extracted")  # Извлекаем файл
        workbook = load_workbook(extracted_xlsx_path)  # Загружаем файл
        sheet = workbook.active
        xlsx_rows = [list(row) for row in sheet.iter_rows(values_only=True)]  # Преобразуем строки в списки
        expected_xlsx_data = [["name", "age", "city"], ["Anna", 25, "Saint-Petersburg"], ["Ivan", 30, "Moscow"]]
        assert xlsx_rows == expected_xlsx_data, f"Содержимое XLSX файла некорректно: {xlsx_rows} != {expected_xlsx_data}"
        assert len(xlsx_rows) > 2, "XLSX файл содержит недостаточно строк."


def test_clean_after_tests():
    # Тест проверки чистоты после тестов.
    cleanup([], "extracted")
    assert not os.path.exists("extracted"), "Директория 'extracted' не была удалена."
